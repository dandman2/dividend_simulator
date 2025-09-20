import yfinance as yf
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os

home_folder = str((Path.home() / "Downloads" / "Dividend_Simulator").resolve())
os.makedirs(home_folder, exist_ok=True)

def excel_serial_date(dt):
    """Convert datetime to Excel serial date."""
    if dt.tzinfo:
        dt = dt.replace(tzinfo=None)
    return (dt - datetime(1899, 12, 30)).days + (dt.second / 86400)

def fetch_dividend_data(ticker_symbol, start_date, end_date):
    """Fetch dividend and price data."""
    ticker = yf.Ticker(ticker_symbol)
    dividends = ticker.dividends[start_date:end_date]
    prices = ticker.history(start=start_date, end=end_date)['Close']
    return dividends, prices

def get_price(ticker, date):
    """Fetch closing price for a date."""
    end_date = (datetime.strptime(date, "%Y-%m-%d") + timedelta(days=3)).strftime("%Y-%m-%d")
    price = yf.download(ticker, start=date, end=end_date, auto_adjust=False)['Close'].iloc[0]
    return float(str(price).split()[2])

def generate_dividend_excel(ticker, start_date, end_date, output_file, shares=1000, exchange_rate=3.69, tax_rate=0.25):
    """Generate Excel file with dividend data."""
    dividends, prices = fetch_dividend_data(ticker, start_date, end_date)
    
    # Prepare data and calculate numerical values for totals
    data = []
    total_received = 0
    total_minus_tax = 0
    total_nis_d_tax = 0
    for date, amount in dividends.items():
        nearest_idx = prices.index.get_indexer([date], method='nearest')[0]
        price = prices.iloc[nearest_idx]
        received = amount * shares
        minus_tax = received * (1 - tax_rate)
        nis_d_tax = minus_tax * exchange_rate
        total_received += received
        total_minus_tax += minus_tax
        total_nis_d_tax += nis_d_tax
        data.append({
            'Date': excel_serial_date(date),
            'Share Value': price,
            'Dividend': amount,
            'Shares': shares,
            'Worth': "=B{row}*D{row}",  # Formula: Share Value * Shares
            'NIS (Worth)': "=E{row}*" + str(exchange_rate),  # Formula: Worth * Exchange Rate
            'Received (D)': "=C{row}*D{row}",  # Formula: Dividend * Shares
            'Minus Tax': "=G{row}*(1-" + str(tax_rate) + ")",  # Formula: Received * (1 - Tax Rate)
            'NIS (D - Tax)': "=H{row}*" + str(exchange_rate)  # Formula: Minus Tax * Exchange Rate
        })

    # Add start/end rows
    start_price = get_price(ticker, start_date)
    end_price = get_price(ticker, end_date)
    for date, price in [(start_date, start_price), (end_date, end_price)]:
        data.insert(0 if date == start_date else len(data), {
            'Date': excel_serial_date(datetime.strptime(date, "%Y-%m-%d")),
            'Share Value': price,
            'Dividend': 0,
            'Shares': shares,
            'Worth': "=B{row}*D{row}",
            'NIS (Worth)': "=E{row}*" + str(exchange_rate),
            'Received (D)': 0,
            'Minus Tax': 0,
            'NIS (D - Tax)': 0
        })

    df = pd.DataFrame(data)
    if df.empty:
        return

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Dividend Simulator"

    # Define styles
    styles = {
        'number': NamedStyle(name='number', number_format='#,##0.00'),
        'no_digits': NamedStyle(name='no_digits', number_format='#,##0'),
        'dollar': NamedStyle(name='dollar', number_format='$ #,##0.00'),
        'dollar_no_digits': NamedStyle(name='dollar_no_digits', number_format='$ #,##0'),
        'nis': NamedStyle(name='nis', number_format='₪ #,##0'),
        'percent': NamedStyle(name='percent', number_format='0.00%'),
        'date': NamedStyle(name='date', number_format='DD/MM/YYYY')
    }

    # Write headers
    headers = ['Date', 'Share Value', 'Dividend', 'Shares', 'Worth', 'NIS (Worth)', 'Received (D)', 'Minus Tax', 'NIS (D - Tax)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.value, cell.font, cell.alignment = header, Font(bold=True), Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")

    # Write data
    for row, record in enumerate(df.itertuples(), 2):
        ws.cell(row, 1).value, ws.cell(row, 1).style = record.Date, styles['date']
        ws.cell(row, 1).font, ws.cell(row, 1).fill = Font(bold=True), PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")
        
        for col, (value, style) in enumerate([
            (record._2, 'dollar'), (record.Dividend if record.Dividend else '', 'number'),
            (record.Shares, 'no_digits'), 
            (record.Worth.format(row=row), 'dollar_no_digits'),  # Use formula
            (record._6.format(row=row), 'nis'),  # Use formula
            (record._7 if isinstance(record._7, (int, float)) else record._7.format(row=row), 'dollar_no_digits'),  # Use formula or 0
            (record._8 if isinstance(record._8, (int, float)) else record._8.format(row=row), 'dollar_no_digits'),  # Use formula or 0
            (record._9 if isinstance(record._9, (int, float)) else record._9.format(row=row), 'nis')  # Use formula or 0
        ], 2):
            ws.cell(row, col).value, ws.cell(row, col).style = value, styles[style]

    # Calculate totals
    totals = {
        'dividend': df['Dividend'].sum(),
        'received': total_received,
        'minus_tax': total_minus_tax,
        'nis_d_tax': total_nis_d_tax,
        'spend': df['Share Value'].iloc[0] * shares,
        'final_worth': df['Share Value'].iloc[-1] * shares
    }
    totals['worth'] = totals['final_worth'] + totals['minus_tax']
    totals['profit'] = totals['worth'] - totals['spend']
    totals['profit_percent'] = totals['profit'] / totals['spend'] if totals['spend'] else 0

    # Write total row
    last_row = len(df) + 1
    row = len(df) + 2
    ws.cell(row, 1).value, ws.cell(row, 1).font = "Total", Font(size=14, bold=True)
    for col, (value, style, fill) in enumerate([
        (df['Share Value'].iloc[-1], 'dollar', None),
        (f"=SUM(C2:C{last_row})", 'number', None),
        (shares, 'no_digits', None),
        (totals['final_worth'], 'dollar_no_digits', "FFFF00"),
        (totals['final_worth'] * exchange_rate, 'nis', None),
        (f"=SUM(G2:G{last_row})", 'dollar_no_digits', None),
        (f"=SUM(H2:H{last_row})", 'dollar_no_digits', "FFFF00"),
        (f"=SUM(I2:I{last_row})", 'nis', None)
    ], 2):
        cell = ws.cell(row, col)
        cell.value, cell.style, cell.font = value, styles[style], Font(size=14, bold=True)
        if fill:
            cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")

    # Write summary
    for r, (label, value, style, fill) in enumerate([
        ("Spend:", totals['spend'], 'dollar_no_digits', "FFC000"),
        ("Total Worth:", totals['worth'], 'dollar_no_digits', "FFFF00"),
        ("Profit:", totals['profit'], 'dollar_no_digits', "92D050")
    ], row + 2):
        ws.cell(r, 5).value, ws.cell(r, 5).alignment, ws.cell(r, 5).font = label, Alignment(horizontal='right'), Font(size=14, bold=True)
        cell = ws.cell(r, 6)
        cell.value, cell.style, cell.font, cell.fill = value, styles[style], Font(size=14, bold=True), PatternFill(start_color=fill, end_color=fill, fill_type="solid")     
    
    ws.cell(row + 2, 7).value, ws.cell(row + 2, 7).style = f"=F{row + 2}*{exchange_rate}", styles['nis']
    ws.cell(row + 2, 7).font, ws.cell(row + 2, 7).fill = Font(size=14, bold=True), PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    ws.cell(row + 3, 7).value, ws.cell(row + 3, 7).style = f"=F{row + 3}*{exchange_rate}", styles['nis']
    ws.cell(row + 3, 7).font, ws.cell(row + 3, 7).fill = Font(size=14, bold=True), PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    ws.cell(row + 4, 7).value, ws.cell(row + 4, 7).style = f"=F{row + 4}*{exchange_rate}", styles['nis']
    ws.cell(row + 4, 7).font, ws.cell(row + 4, 7).fill = Font(size=14, bold=True), PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    ws.cell(row + 4, 8).value, ws.cell(row + 4, 8).style = f"=F{row + 4}/F{row + 2}", styles['percent']
    ws.cell(row + 4, 8).font, ws.cell(row + 4, 8).fill = Font(size=14, bold=True), PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    ws.cell(2, 5).fill = ws.cell(2, 6).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    # Format columns and borders
    thin, thick = Side(style="thin"), Side(style="medium")
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15
        for cell in ws[get_column_letter(col)][1:row]:
            cell.alignment = Alignment(horizontal='right')
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for cell in ws[row][0:9]:
        cell.border = Border(top=thick, bottom=thick)
    ws[row][0].border = Border(top=thick, bottom=thick, left=thick)
    ws[row][8].border = Border(top=thick, bottom=thick, right=thick)

    for row in ws[f'E{row+2}:G{row+4}']:
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb.save(output_file)
    # Open the file
    os.startfile(output_file)

if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python generate_dividend_excel.py <ticker> <start_date> <end_date> [<shares> <exchange_rate> <tax_rate>]\n"
              "Example: python generate_dividend_excel.py AAPL 01.01.2023 31.12.2024 1000 3.69 0.25")
        sys.exit(1)

    ticker, start_date_input, end_date_input = sys.argv[1:4]
    
    # Set default values
    shares = 1000
    exchange_rate = 3.69
    tax_rate = 0.25
    
    # Parse optional parameters
    try:
        if len(sys.argv) > 4:
            shares = int(sys.argv[4])
        if len(sys.argv) > 5:
            exchange_rate = float(sys.argv[5])
        if len(sys.argv) > 6:
            tax_rate = float(sys.argv[6])
    except ValueError:
        print("Error: Shares must be an integer, exchange_rate and tax_rate must be floats")
        sys.exit(1)

    try:
        start_date = datetime.strptime(start_date_input, "%d.%m.%Y").strftime("%Y-%m-%d")
        end_date = datetime.strptime(end_date_input, "%d.%m.%Y").strftime("%Y-%m-%d")
        output_path = f"{home_folder}\\{ticker}_Sim.xlsx"
        generate_dividend_excel(ticker, start_date, end_date, output_path, shares, exchange_rate, tax_rate)
    except ValueError:
        print("Error: Invalid date format. Use DD.MM.YYYY (e.g., 01.01.2023)")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)